from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse, JSONResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, and_
from database import get_db
from core.security import get_current_user
from core.permissions import require_worker, require_manager
from core.audit import log_action
from models.user import User
from models.zone import Zone
from models.cell import Cell
from models.product import Product
from models.order import Order, OrderItem
from models.stock import Stock
from typing import List, Optional
from models.document import WarehouseDocument, DocStatus, DocumentItem
from models.product_image import ProductImage
from schemas.catalog import (
    ZoneCreate, ZoneResponse,
    CellCreate, CellResponse,
    ProductCreate, ProductUpdate, ProductResponse
)
from routers.notifications import create_notification
import csv, io, json
from datetime import datetime
import os
import uuid
import shutil

try:
    import openpyxl
    from openpyxl import Workbook
    XLSX_SUPPORTED = True
except ImportError:
    XLSX_SUPPORTED = False

router = APIRouter(prefix="/api/catalog", tags=["Справочники склада"])

# === ЗОНЫ ===
@router.post("/zones", response_model=ZoneResponse)
def create_zone(data: ZoneCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    zone = Zone(**data.model_dump())
    db.add(zone)
    db.commit()
    db.refresh(zone)
    log_action(db, current_user, "CREATE", "zone", zone.id, new_value={"code": zone.code})
    
    managers = db.query(User).filter(User.role.in_(['admin', 'warehouse_manager'])).all()
    for manager in managers:
        create_notification(
            db=db,
            user_id=manager.id,
            type="system",
            title="📦 Новая зона",
            message=f"Создана зона '{zone.code}'",
            link="/products"
        )
    return zone

@router.get("/zones", response_model=list[ZoneResponse])
def list_zones(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    return db.query(Zone).filter(Zone.is_active == True).all()

# === ЯЧЕЙКИ ===
@router.post("/cells", response_model=CellResponse)
def create_cell(data: CellCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    cell = Cell(**data.model_dump())
    db.add(cell)
    db.commit()
    db.refresh(cell)
    log_action(db, current_user, "CREATE", "cell", cell.id, new_value={"code": cell.code})
    
    managers = db.query(User).filter(User.role.in_(['admin', 'warehouse_manager'])).all()
    for manager in managers:
        create_notification(
            db=db,
            user_id=manager.id,
            type="system",
            title="📦 Новая ячейка",
            message=f"Создана ячейка '{cell.code}'",
            link="/products"
        )
    return cell

@router.get("/cells", response_model=list[CellResponse])
def list_cells(zone_id: int = Query(None), db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    query = db.query(Cell)
    if zone_id:
        query = query.filter(Cell.zone_id == zone_id)
    return query.all()

# === ТОВАРЫ ===
@router.post("/products", response_model=ProductResponse)
def create_product(data: ProductCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    try:
        product = Product(
            sku=data.sku, name=data.name, category=data.category, description=data.description,
            unit=data.unit, weight_kg=data.weight_kg, volume_m3=data.volume_m3, barcode=data.barcode,
            purchase_price=data.purchase_price, sale_price=data.sale_price, min_stock=data.min_stock, max_stock=data.max_stock
        )
        db.add(product)
        db.flush()
        
        log_action(db, current_user, "CREATE", "product", product.id, 
                   new_value={"sku": product.sku, "name": product.name, "category": product.category})
        
        db.commit()
        db.refresh(product)
        
        managers = db.query(User).filter(User.role.in_(['admin', 'warehouse_manager'])).all()
        for manager in managers:
            create_notification(
                db=db,
                user_id=manager.id,
                type="system",
                title="📦 Новый товар",
                message=f"Добавлен товар '{product.name}' (SKU: {product.sku})",
                link="/products"
            )
        return product
    except Exception as e:
        db.rollback()
        raise HTTPException(500, str(e))

@router.get("/products")
def list_products(search: str = Query(None), category: str = Query(None), db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    query = db.query(
        Product,
        func.coalesce(func.sum(Stock.quantity), 0).label('total_qty')
    ).outerjoin(
        Stock, Stock.product_id == Product.id
    ).filter(
        Product.is_active == True
    ).group_by(Product.id)

    if search:
        query = query.filter(Product.name.ilike(f"%{search}%") | Product.sku.ilike(f"%{search}%") | Product.category.ilike(f"%{search}%"))
    if category:
        query = query.filter(Product.category == category)

    results = query.all()
    
    products_list = []
    for product, total_qty in results:
        p_dict = {
            "id": product.id,
            "sku": product.sku,
            "name": product.name,
            "category": product.category,
            "description": product.description,
            "unit": product.unit,
            "weight_kg": float(product.weight_kg or 0),
            "volume_m3": float(product.volume_m3 or 0),
            "barcode": product.barcode,
            "purchase_price": float(product.purchase_price or 0),
            "sale_price": float(product.sale_price or 0),
            "min_stock": product.min_stock,
            "max_stock": product.max_stock,
            "is_active": product.is_active,
            "quantity": total_qty, 
            "total_quantity": total_qty
        }
        products_list.append(p_dict)
        
    return products_list

@router.get("/products/export")
def export_products(
    format: str = Query("csv", description="Формат экспорта: csv или xlsx"),
    db: Session = Depends(get_db), 
    current_user: User = Depends(get_current_user)
):
    products = db.query(Product).filter(Product.is_active == True).all()
    
    log_action(db, current_user, "EXPORT", "product", 0, 
               new_value={"format": format, "count": len(products)})
    
    if format.lower() == "xlsx":
        if not XLSX_SUPPORTED:
            raise HTTPException(500, "Библиотека openpyxl не установлена")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Товары"
        
        headers = ["SKU", "Название", "Категория", "Вес (кг)", "Мин. остаток", "Макс. остаток", "Закупка (₽)", "Продажа (₽)"]
        ws.append(headers)
        
        for p in products:
            ws.append([
                p.sku or "", p.name or "", p.category or "", p.weight_kg or 0,
                p.min_stock or 0, p.max_stock or 0,
                float(p.purchase_price or 0), float(p.sale_price or 0)
            ])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"products_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    else:
        output = io.StringIO()
        writer = csv.writer(output, delimiter=',', quoting=csv.QUOTE_MINIMAL)
        
        writer.writerow(["SKU", "Название", "Категория", "Вес (кг)", "Мин. остаток", "Макс. остаток", "Закупка (₽)", "Продажа (₽)"])
        
        for p in products:
            writer.writerow([
                p.sku or "", p.name or "", p.category or "", p.weight_kg or 0,
                p.min_stock or 0, p.max_stock or 0,
                float(p.purchase_price or 0), float(p.sale_price or 0)
            ])
        
        output.seek(0)
        
        b_output = io.BytesIO()
        b_output.write(output.getvalue().encode('utf-8-sig'))
        b_output.seek(0)
        
        filename = f"products_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        return StreamingResponse(
            b_output,
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

@router.put("/products/{product_id}", response_model=ProductResponse)
def update_product(product_id: int, data: ProductUpdate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    try:
        product = db.query(Product).filter(Product.id == product_id, Product.is_active == True).first()
        if not product: raise HTTPException(404, detail="Товар не найден")
        
        old_value = {"name": product.name, "sku": product.sku, "category": product.category, "price": float(product.sale_price or 0)}
        
        update_data = data.model_dump(exclude_unset=True)
        for key, value in update_data.items():
            setattr(product, key, value)
            
        db.commit()
        db.refresh(product)
        
        log_action(db, current_user, "UPDATE", "product", product_id, 
                   old_value=old_value, 
                   new_value={"name": product.name, "sku": product.sku, "category": product.category, "price": float(product.sale_price or 0)})
        
        managers = db.query(User).filter(User.role.in_(['admin', 'warehouse_manager'])).all()
        for manager in managers:
            create_notification(
                db=db,
                user_id=manager.id,
                type="system",
                title="💰 Изменён товар",
                message=f"Обновлён товар '{product.name}' (SKU: {product.sku})",
                link="/products"
            )
        return product
    except Exception as e:
        db.rollback()
        raise HTTPException(500, str(e))

@router.delete("/products/{product_id}")
def delete_product(
    product_id: int, 
    hard: bool = Query(False), 
    db: Session = Depends(get_db), 
    current_user: User = Depends(get_current_user)
):
    try:
        product = db.query(Product).filter(Product.id == product_id, Product.is_active == True).first()
        if not product: 
            raise HTTPException(404, detail="Товар не найден")
        
        if hard:
            has_deps = db.query(OrderItem.id).filter(OrderItem.product_id == product_id).first() or \
                       db.query(Stock.id).filter(Stock.product_id == product_id, Stock.quantity > 0).first() or \
                       db.query(DocumentItem.id).filter(DocumentItem.product_id == product_id).first()
            if has_deps: 
                raise HTTPException(400, detail="Есть зависимости. Используйте архивацию.")
            
            db.delete(product)
            db.commit()
            log_action(db, current_user, "HARD_DELETE", "product", product_id, 
                       new_value={"sku": product.sku, "name": product.name})
            
            managers = db.query(User).filter(User.role.in_(['admin', 'warehouse_manager'])).all()
            for manager in managers:
                create_notification(
                    db=db,
                    user_id=manager.id,
                    type="system",
                    title="🗑️ Товар удалён",
                    message=f"Товар '{product.name}' (SKU: {product.sku}) полностью удалён",
                    link="/archive"
                )
            return {"message": "Товар полностью удалён", "hard_delete": True}
        else:
            product.is_active = False
            product.archived_at = datetime.now()
            db.commit()
            log_action(db, current_user, "ARCHIVE", "product", product_id, 
                       new_value={"sku": product.sku, "name": product.name, "archived_at": str(product.archived_at)})
            
            managers = db.query(User).filter(User.role.in_(['admin', 'warehouse_manager'])).all()
            for manager in managers:
                create_notification(
                    db=db,
                    user_id=manager.id,
                    type="system",
                    title="🗑️ Товар в архиве",
                    message=f"Товар '{product.name}' (SKU: {product.sku}) перемещён в архив",
                    link="/archive"
                )
            return {"message": "Товар архивирован", "hard_delete": False}
    except HTTPException: 
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(500, str(e))

# === АРХИВ ===
@router.get("/products/archived")
def list_archived_products(
    search: str = Query(None),
    date_from: str = Query(None),
    date_to: str = Query(None),
    db: Session = Depends(get_db), 
    _: User = Depends(get_current_user)
):
    query = db.query(Product).filter(Product.is_active == False)
    
    if search:
        query = query.filter(
            Product.name.ilike(f"%{search}%") | 
            Product.sku.ilike(f"%{search}%")
        )
    
    if date_from:
        query = query.filter(Product.archived_at >= datetime.fromisoformat(date_from))
    if date_to:
        query = query.filter(Product.archived_at <= datetime.fromisoformat(date_to))
    
    query = query.order_by(Product.archived_at.desc())
    
    results = []
    for p in query.all():
        results.append({
            "id": p.id,
            "sku": p.sku,
            "name": p.name,
            "category": p.category,
            "description": p.description,
            "unit": p.unit,
            "weight_kg": float(p.weight_kg or 0),
            "volume_m3": float(p.volume_m3 or 0),
            "barcode": p.barcode,
            "purchase_price": float(p.purchase_price or 0),
            "sale_price": float(p.sale_price or 0),
            "min_stock": p.min_stock,
            "max_stock": p.max_stock,
            "is_active": p.is_active,
            "archived_at": p.archived_at.isoformat() if p.archived_at else None
        })
    
    return results

@router.post("/products/{product_id}/restore")
def restore_product(
    product_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    product = db.query(Product).filter(
        Product.id == product_id, 
        Product.is_active == False
    ).first()
    
    if not product:
        raise HTTPException(404, detail="Товар не найден в архиве")
    
    product.is_active = True
    product.archived_at = None
    db.commit()
    
    log_action(db, current_user, "RESTORE", "product", product_id, 
               new_value={"sku": product.sku, "name": product.name})
    
    managers = db.query(User).filter(User.role.in_(['admin', 'warehouse_manager'])).all()
    for manager in managers:
        create_notification(
            db=db,
            user_id=manager.id,
            type="system",
            title="♻️ Товар восстановлен",
            message=f"Товар '{product.name}' (SKU: {product.sku}) восстановлен из архива",
            link="/products"
        )
    
    return {"message": f"Товар '{product.name}' восстановлен"}

@router.delete("/products/{product_id}/permanent")
def delete_product_permanent(
    product_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Полное удаление товара из архива"""
    try:
        product = db.query(Product).filter(Product.id == product_id).first()
        if not product:
            raise HTTPException(404, detail="Товар не найден")
        
        has_deps = db.query(OrderItem.id).filter(OrderItem.product_id == product_id).first() or \
                   db.query(Stock.id).filter(Stock.product_id == product_id, Stock.quantity > 0).first() or \
                   db.query(DocumentItem.id).filter(DocumentItem.product_id == product_id).first()
        
        if has_deps:
            raise HTTPException(400, detail="Есть зависимости (заказы, остатки, документы). Нельзя удалить.")
        
        db.delete(product)
        db.commit()
        
        log_action(db, current_user, "PERMANENT_DELETE", "product", product_id,
                   new_value={"sku": product.sku, "name": product.name})
        
        managers = db.query(User).filter(User.role.in_(['admin', 'warehouse_manager'])).all()
        for manager in managers:
            create_notification(
                db=db,
                user_id=manager.id,
                type="system",
                title="🗑️ Товар полностью удалён",
                message=f"Товар '{product.name}' (SKU: {product.sku}) полностью удалён из системы",
                link="/archive"
            )
        
        return {"message": "Товар полностью удалён"}
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(500, str(e))

@router.post("/products/bulk-delete")
def bulk_delete_products(
    request_data: dict,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if current_user.role not in ["admin", "warehouse_manager"]:
        raise HTTPException(403, "Доступ запрещён. Требуются права менеджера.")
    
    product_ids = request_data.get("product_ids", [])
    hard = request_data.get("hard", False)
    if not product_ids: 
        raise HTTPException(400, "Не выбраны товары")
    
    success, errors, error_details = 0, 0, []
    for pid in product_ids:
        try:
            p = db.query(Product).filter(
                Product.id == pid, 
                Product.is_active == True
            ).first()
            if not p:
                errors += 1
                error_details.append(f"Товар #{pid} не найден")
                continue
            
            if hard:
                has_deps = db.query(OrderItem.id).filter(OrderItem.product_id == pid).first() or \
                          db.query(Stock.id).filter(Stock.product_id == pid, Stock.quantity > 0).first()
                if has_deps: 
                    errors += 1
                    error_details.append(f"Товар #{pid}: есть зависимости")
                    continue
                db.delete(p)
            else:
                p.is_active = False
                p.archived_at = datetime.now()
            
            log_action(db, current_user, "HARD_DELETE" if hard else "ARCHIVE", "product", pid, 
                       new_value={"sku": p.sku})
            success += 1
        except Exception as e:
            errors += 1
            error_details.append(f"Товар #{pid}: {str(e)}")
            
    db.commit()
    
    if success > 0:
        managers = db.query(User).filter(User.role.in_(['admin', 'warehouse_manager'])).all()
        for manager in managers:
            create_notification(
                db=db,
                user_id=manager.id,
                type="system",
                title="🗑️ Массовое удаление",
                message=f"Удалено {success} товаров" + (" (полностью)" if hard else " (в архив)"),
                link="/archive"
            )
    
    return {
        "success": success,
        "errors": errors,
        "error_details": error_details[:10],
        "hard_delete": hard
    }

@router.post("/products/import")
async def import_products(file: UploadFile = File(...), db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    if current_user.role not in ["admin", "warehouse_manager"]:
        raise HTTPException(403, "Доступ запрещён. Требуются права менеджера.")
    
    if not file.filename: return JSONResponse(status_code=400, content={"status": "error", "message": "Файл не выбран"})
    try:
        if file.filename.lower().endswith('.csv'): result = await import_from_csv(file, db)
        elif file.filename.lower().endswith(('.xlsx', '.xls')): result = await import_from_xlsx(file, db)
        else: return JSONResponse(status_code=400, content={"status": "error", "message": "Неподдерживаемый формат"})
        
        if result.get("status") == "success":
            log_action(db, current_user, "BULK_IMPORT", "product", 0, 
                       new_value={"file": file.filename, "created": result["created"], "updated": result["updated"]})
            
            managers = db.query(User).filter(User.role.in_(['admin', 'warehouse_manager'])).all()
            for manager in managers:
                create_notification(
                    db=db,
                    user_id=manager.id,
                    type="system",
                    title="📥 Импорт товаров",
                    message=f"Импортировано {result['created']} новых и обновлено {result['updated']} товаров",
                    link="/products"
                )
        return result
    except Exception as e:
        return JSONResponse(status_code=500, content={"status": "error", "message": str(e)})

async def import_from_csv(file: UploadFile, db: Session):
    content = await file.read()
    for enc in ['utf-8-sig', 'utf-8', 'cp1251']:
        try: text = content.decode(enc); break
        except: continue
    else: return {"status": "errors", "created": 0, "updated": 0, "errors": ["Неверная кодировка"]}
    
    delimiter = ',' if ',' in text.split('\n')[0] else ';'
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    header_map = {'sku':'sku','название':'name','name':'name','категория':'category','category':'category',
                  'вес (кг)':'weight_kg','weight_kg':'weight_kg','мин. остаток':'min_stock','min_stock':'min_stock',
                  'макс. остаток':'max_stock','max_stock':'max_stock','закупка (₽)':'purchase_price','purchase_price':'purchase_price',
                  'продажа (₽)':'sale_price','sale_price':'sale_price'}
    
    created, updated, errors = 0, 0, []
    for i, row in enumerate(reader, 2):
        try:
            norm = {header_map.get(k.strip().lower(), k.strip().lower()): v for k, v in row.items() if k}
            sku = norm.get("sku", "").strip()
            if not sku: raise ValueError("Нет SKU")
            
            existing = db.query(Product).filter(Product.sku == sku).first()
            def sf(v, d=0): return float(str(v).replace(',','.').strip()) if v else d
            def si(v, d=0): return int(float(str(v).replace(',','.').strip())) if v else d
            
            data = {"name": norm.get("name","Без имени").strip(), "category": norm.get("category") or None,
                    "weight_kg": sf(norm.get("weight_kg")), "min_stock": si(norm.get("min_stock")),
                    "max_stock": si(norm.get("max_stock"), 100), "purchase_price": sf(norm.get("purchase_price")),
                    "sale_price": sf(norm.get("sale_price"))}
            
            if existing:
                for k,v in data.items(): 
                    if v is not None: setattr(existing, k, v)
                updated += 1
            else:
                db.add(Product(sku=sku, **data))
                db.flush()
                created += 1
        except Exception as e: errors.append(f"Строка {i}: {e}")
            
    if not errors: db.commit(); return {"status": "success", "created": created, "updated": updated}
    db.rollback(); return {"status": "errors", "created": created, "updated": updated, "errors": errors[:10]}

async def import_from_xlsx(file: UploadFile, db: Session):
    if not XLSX_SUPPORTED: return {"status": "errors", "created": 0, "updated": 0, "errors": ["openpyxl не установлен"]}
    try: wb = openpyxl.load_workbook(io.BytesIO(await file.read()))
    except Exception as e: return {"status": "errors", "created": 0, "updated": 0, "errors": [str(e)]}
    
    ws = wb.active; created, updated, errors = 0, 0, []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        try:
            if not row[0]: continue
            sku = str(row[0]).strip()
            existing = db.query(Product).filter(Product.sku == sku).first()
            def sf(v, d=0): return float(str(v).replace(',','.').strip()) if v is not None else d
            def si(v, d=0): return int(float(str(v).replace(',','.').strip())) if v is not None else d
            
            data = {"name": str(row[1]).strip() if row[1] else "Без имени", "category": str(row[2]).strip() if row[2] else None,
                    "weight_kg": sf(row[3]), "min_stock": si(row[4]), "max_stock": si(row[5], 100),
                    "purchase_price": sf(row[6]), "sale_price": sf(row[7])}
            
            if existing:
                for k,v in data.items(): 
                    if v is not None: setattr(existing, k, v)
                updated += 1
            else:
                db.add(Product(sku=sku, **data))
                db.flush()
                created += 1
        except Exception as e: errors.append(f"Строка {idx}: {e}")
            
    if not errors: db.commit(); return {"status": "success", "created": created, "updated": updated}
    db.rollback(); return {"status": "errors", "created": created, "updated": updated, "errors": errors[:10]}

# === КАТЕГОРИИ ===
@router.get("/categories")
def get_categories(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    results = db.query(Product.category, func.count(Product.id).label('cnt')).filter(
        Product.is_active == True, Product.category.isnot(None), Product.category != ''
    ).group_by(Product.category).order_by(Product.category).all()
    return [{"name": r[0], "product_count": r[1]} for r in results if r[0]]

@router.post("/categories", status_code=201)
def create_category(name: str = Query(...), db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    if current_user.role not in ["admin", "warehouse_manager"]:
        raise HTTPException(403, "Доступ запрещён. Требуются права менеджера.")
    
    name = name.strip()
    if db.query(Product.id).filter(Product.category.ilike(name), Product.is_active == True).first():
        raise HTTPException(400, detail="Категория существует")
    log_action(db, current_user, "CREATE", "category", 0, new_value={"name": name})
    
    managers = db.query(User).filter(User.role.in_(['admin', 'warehouse_manager'])).all()
    for manager in managers:
        create_notification(
            db=db,
            user_id=manager.id,
            type="system",
            title="📂 Новая категория",
            message=f"Создана категория '{name}'",
            link="/products"
        )
    return {"name": name, "message": "Готово"}

@router.put("/categories/{old_name}")
def update_category(old_name: str, new_name: str = Query(...), db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    if current_user.role not in ["admin", "warehouse_manager"]:
        raise HTTPException(403, "Доступ запрещён. Требуются права менеджера.")
    
    old_name, new_name = old_name.strip(), new_name.strip()
    if old_name == new_name: raise HTTPException(400, detail="Имена совпадают")
    if db.query(Product.id).filter(Product.category.ilike(new_name), Product.is_active == True).first():
        raise HTTPException(400, detail="Новая категория существует")
    
    products = db.query(Product).filter(Product.category == old_name, Product.is_active == True).all()
    if not products: raise HTTPException(404, detail="Категория пуста")
    
    cnt = db.query(Product).filter(Product.category == old_name, Product.is_active == True).update({"category": new_name}, synchronize_session=False)
    db.commit()
    log_action(db, current_user, "UPDATE", "category", 0, old_value={"name": old_name}, new_value={"name": new_name, "affected": cnt})
    
    managers = db.query(User).filter(User.role.in_(['admin', 'warehouse_manager'])).all()
    for manager in managers:
        create_notification(
            db=db,
            user_id=manager.id,
            type="system",
            title="📂 Категория переименована",
            message=f"Категория '{old_name}' переименована в '{new_name}'",
            link="/products"
        )
    return {"message": "Обновлено", "updated_products": cnt}

@router.delete("/categories/{name}")
def delete_category(name: str, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    if current_user.role not in ["admin", "warehouse_manager"]:
        raise HTTPException(403, "Доступ запрещён. Требуются права менеджера.")
    
    name = name.strip()
    products = db.query(Product).filter(Product.category == name, Product.is_active == True).all()
    if not products: raise HTTPException(404, detail="Категория пуста")
    
    cnt = db.query(Product).filter(Product.category == name, Product.is_active == True).update({"category": None}, synchronize_session=False)
    db.commit()
    log_action(db, current_user, "DELETE", "category", 0, new_value={"name": name, "affected": cnt})
    
    managers = db.query(User).filter(User.role.in_(['admin', 'warehouse_manager'])).all()
    for manager in managers:
        create_notification(
            db=db,
            user_id=manager.id,
            type="system",
            title="📂 Категория удалена",
            message=f"Категория '{name}' удалена, {cnt} товаров без категории",
            link="/products"
        )
    return {"message": f"Удалена. {cnt} товаров без категории."}


# ============================================
# 🔥 ФОТО ТОВАРОВ
# ============================================

@router.post("/products/{product_id}/images")
async def upload_product_images(
    product_id: int,
    files: List[UploadFile] = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Загрузить фото для товара"""
    if current_user.role not in ["admin", "warehouse_manager"]:
        raise HTTPException(403, "Доступ запрещён")
    
    product = db.query(Product).filter(Product.id == product_id).first()
    if not product:
        raise HTTPException(404, "Товар не найден")
    
    upload_dir = "static/products"
    os.makedirs(upload_dir, exist_ok=True)
    
    uploaded_images = []
    for file in files:
        ext = os.path.splitext(file.filename)[1]
        filename = f"product_{product_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:6]}{ext}"
        filepath = os.path.join(upload_dir, filename)
        
        with open(filepath, "wb") as f:
            content = await file.read()
            f.write(content)
        
        existing_main = db.query(ProductImage).filter(
            ProductImage.product_id == product_id,
            ProductImage.is_main == True
        ).first()
        
        image = ProductImage(
            product_id=product_id,
            image_url=f"/static/products/{filename}",
            is_main=not bool(existing_main),
            order=len(db.query(ProductImage).filter(ProductImage.product_id == product_id).all())
        )
        db.add(image)
        uploaded_images.append(image)
    
    db.commit()
    return {"message": f"Загружено {len(uploaded_images)} фото", "images": uploaded_images}


@router.get("/products/{product_id}/images")
def get_product_images(
    product_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Получить фото товара"""
    images = db.query(ProductImage).filter(
        ProductImage.product_id == product_id
    ).order_by(ProductImage.is_main.desc(), ProductImage.order.asc()).all()
    
    return [
        {
            "id": img.id,
            "image_url": img.image_url,
            "is_main": img.is_main,
            "order": img.order
        }
        for img in images
    ]


@router.delete("/products/images/{image_id}")
def delete_product_image(
    image_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Удалить фото товара"""
    if current_user.role not in ["admin", "warehouse_manager"]:
        raise HTTPException(403, "Доступ запрещён")
    
    image = db.query(ProductImage).filter(ProductImage.id == image_id).first()
    if not image:
        raise HTTPException(404, "Фото не найдено")
    
    file_path = image.image_url.lstrip('/')
    if os.path.exists(file_path):
        os.remove(file_path)
    
    if image.is_main:
        next_image = db.query(ProductImage).filter(
            ProductImage.product_id == image.product_id,
            ProductImage.id != image_id
        ).order_by(ProductImage.order.asc()).first()
        if next_image:
            next_image.is_main = True
    
    db.delete(image)
    db.commit()
    
    return {"message": "Фото удалено"}


@router.post("/products/images/{image_id}/set-main")
def set_main_image(
    image_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Сделать фото главным"""
    if current_user.role not in ["admin", "warehouse_manager"]:
        raise HTTPException(403, "Доступ запрещён")
    
    image = db.query(ProductImage).filter(ProductImage.id == image_id).first()
    if not image:
        raise HTTPException(404, "Фото не найдено")
    
    db.query(ProductImage).filter(
        ProductImage.product_id == image.product_id
    ).update({"is_main": False})
    
    image.is_main = True
    db.commit()
    
    return {"message": "Главное фото установлено"}