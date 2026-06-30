from fastapi import APIRouter, Depends, HTTPException, status, Response
from sqlalchemy.orm import Session, joinedload
from sqlalchemy.exc import IntegrityError
from database import get_db
from core.permissions import require_manager, require_worker
from core.audit import log_action
from models.user import User
from models.document import WarehouseDocument, DocumentItem, DocStatus, DocType
from models.product import Product
from models.stock import Stock
from models.cell import Cell
from schemas.documents import DocumentCreate, DocumentResponse, DocumentItemCreate
from routers.notifications import create_notification
import uuid
import io
import csv
from datetime import datetime

# 🔥 ИМПОРТЫ ДЛЯ EXCEL
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    XLSX_SUPPORTED = True
except ImportError:
    XLSX_SUPPORTED = False
    print("⚠️ openpyxl не установлен. Excel экспорт будет недоступен.")

router = APIRouter(prefix="/api/documents", tags=["Складские документы"])

def doc_type_label(doc_type):
    """Человеческое название типа документа"""
    labels = {
        'receive': 'Приёмка',
        'ship': 'Отгрузка',
        'transfer': 'Перемещение',
        'adjust': 'Корректировка'
    }
    return labels.get(str(doc_type), str(doc_type))

@router.post("/", response_model=DocumentResponse, status_code=status.HTTP_201_CREATED)
def create_document(data: DocumentCreate, db: Session = Depends(get_db), current_user: User = require_manager):
    try:
        if not data.items:
            raise HTTPException(400, "Добавьте хотя бы одну позицию")
        
        doc_number = data.doc_number.strip() if data.doc_number else f"DOC-{uuid.uuid4().hex[:8].upper()}"
        
        doc = WarehouseDocument(
            doc_number=doc_number,
            type=data.type,
            operator_id=current_user.id,
            comment=data.comment,
            status=DocStatus.DRAFT
        )
        
        db.add(doc)
        db.flush()
        
        for item in data.items:
            db.add(DocumentItem(
                doc_id=doc.id,
                product_id=item.product_id,
                quantity=item.quantity,
                from_cell_id=item.from_cell_id,
                to_cell_id=item.to_cell_id
            ))
        
        log_action(db, current_user, "CREATE", "document", doc.id, 
                   new_value={"doc_number": doc.doc_number, "type": doc.type, "status": "draft"})
        
        db.commit()
        db.refresh(doc)
        
        # Уведомление для всех кладовщиков
        workers = db.query(User).filter(User.role.in_(['admin', 'warehouse_manager', 'warehouse_worker'])).all()
        for worker in workers:
            create_notification(
                db=db,
                user_id=worker.id,
                type="document",
                title=f"📄 Новый документ {doc.doc_number}",
                message=f"Создан документ {doc.doc_number} типа '{doc_type_label(doc.type)}'",
                link="/documents"
            )
        return doc
        
    except IntegrityError:
        db.rollback()
        raise HTTPException(400, "Документ с таким номером уже существует")
    except Exception as e:
        db.rollback()
        raise HTTPException(500, str(e))

@router.get("/{doc_id}")
def get_document_details(doc_id: int, db: Session = Depends(get_db), user: User = require_worker):
    doc = db.query(WarehouseDocument).filter(WarehouseDocument.id == doc_id).first()
    if not doc:
        raise HTTPException(404, "Документ не найден")
    
    doc_items = db.query(DocumentItem).join(Product).filter(
        DocumentItem.doc_id == doc_id
    ).all()
    
    return {
        "id": doc.id,
        "doc_number": doc.doc_number,
        "type": doc.type,
        "status": doc.status,
        "created_at": doc.created_at,
        "comment": doc.comment,
        "items": [
            {
                "id": item.id,
                "product_id": item.product_id,
                "product_name": item.product.name if item.product else f"Товар #{item.product_id}",
                "product_sku": item.product.sku if item.product else None,
                "quantity": item.quantity,
                "from_cell_id": item.from_cell_id,
                "to_cell_id": item.to_cell_id
            }
            for item in doc_items
        ]
    }

@router.get("/", response_model=list[DocumentResponse])
def list_documents(db: Session = Depends(get_db), current_user: User = require_worker):
    return db.query(WarehouseDocument).order_by(WarehouseDocument.created_at.desc()).all()

@router.post("/{doc_id}/complete", response_model=DocumentResponse)
def complete_document(doc_id: int, db: Session = Depends(get_db), current_user: User = require_worker):
    """Проведение документа (только кладовщик или админ/менеджер для корректировки)"""
    try:
        doc = db.query(WarehouseDocument).filter(WarehouseDocument.id == doc_id).first()
        if not doc:
            raise HTTPException(404, "Документ не найден")
        if doc.status != DocStatus.DRAFT:
            raise HTTPException(400, "Можно провести только документ в статусе ЧЕРНОВИК")
        
        # 🔥 ПРОВЕРКА: корректировку может проводить только админ/менеджер
        is_adjust = str(doc.type) == "adjust"
        is_manager = current_user.role in ['admin', 'warehouse_manager']
        is_worker = current_user.role == 'warehouse_worker'
        
        if is_adjust and is_worker and not is_manager:
            raise HTTPException(403, "Корректировку могут проводить только Администратор или Менеджер склада")
        
        items = db.query(DocumentItem).filter(DocumentItem.doc_id == doc_id).all()
        if not items:
            raise HTTPException(400, "Документ пуст")
        
        # Проверяем и обновляем остатки
        for item in items:
            product = db.query(Product).get(item.product_id)
            if not product:
                raise HTTPException(404, f"Товар {item.product_id} не найден")
            
            doc_type = doc.type.value if hasattr(doc.type, 'value') else str(doc.type)
            
            if doc_type == "receive":
                stock = db.query(Stock).filter(
                    Stock.product_id == item.product_id,
                    Stock.cell_id == item.to_cell_id
                ).first()
                
                if not stock:
                    stock = Stock(
                        product_id=item.product_id,
                        cell_id=item.to_cell_id,
                        quantity=0,
                        cost_price=float(product.purchase_price) if product.purchase_price else 0
                    )
                    db.add(stock)
                    db.flush()
                
                stock.quantity += item.quantity
                
            elif doc_type == "ship":
                stock = db.query(Stock).filter(
                    Stock.product_id == item.product_id,
                    Stock.cell_id == item.from_cell_id
                ).first()
                
                if not stock or stock.quantity < item.quantity:
                    available = stock.quantity if stock else 0
                    raise HTTPException(
                        status_code=400, 
                        detail=f"Недостаточно '{product.name}': доступно {available}, нужно {item.quantity}"
                    )
                
                stock.quantity -= item.quantity

            elif doc_type == "transfer":
                if not item.from_cell_id or not item.to_cell_id:
                    raise HTTPException(400, "Для перемещения укажите ячейки 'Откуда' и 'Куда'")
                
                if item.from_cell_id == item.to_cell_id:
                    raise HTTPException(400, "Ячейки 'Откуда' и 'Куда' не могут совпадать")

                stock_from = db.query(Stock).filter(
                    Stock.product_id == item.product_id,
                    Stock.cell_id == item.from_cell_id
                ).first()

                if not stock_from or stock_from.quantity < item.quantity:
                    available = stock_from.quantity if stock_from else 0
                    raise HTTPException(400, f"Недостаточно товара в ячейке: доступно {available}, нужно {item.quantity}")
                
                stock_from.quantity -= item.quantity

                stock_to = db.query(Stock).filter(
                    Stock.product_id == item.product_id,
                    Stock.cell_id == item.to_cell_id
                ).first()

                if not stock_to:
                    stock_to = Stock(
                        product_id=item.product_id,
                        cell_id=item.to_cell_id,
                        quantity=0
                    )
                    db.add(stock_to)
                    db.flush()
                
                stock_to.quantity += item.quantity
                
            elif doc_type == "adjust":
                stock = db.query(Stock).filter(
                    Stock.product_id == item.product_id,
                    Stock.cell_id == item.to_cell_id
                ).first()

                if item.quantity > 0:
                    if not stock:
                        stock = Stock(
                            product_id=item.product_id,
                            cell_id=item.to_cell_id,
                            quantity=0
                        )
                        db.add(stock)
                        db.flush()
                    stock.quantity += item.quantity
                    
                elif item.quantity < 0:
                    if not stock or stock.quantity < abs(item.quantity):
                        available = stock.quantity if stock else 0
                        raise HTTPException(400, f"Недостаточно товара для списания: доступно {available}, нужно списать {abs(item.quantity)}")
                    stock.quantity += item.quantity
                else:
                    raise HTTPException(400, "Количество не может быть 0")
        
        doc.status = DocStatus.COMPLETED
        db.commit()
        db.refresh(doc)
        
        log_action(
            db=db,
            user=current_user,
            action="COMPLETE",
            entity_type="document",
            entity_id=doc_id,
            old_value={"status": "draft"},
            new_value={"status": "completed"}
        )
        
        # Уведомление для админов и менеджеров
        managers = db.query(User).filter(User.role.in_(['admin', 'warehouse_manager'])).all()
        for manager in managers:
            create_notification(
                db=db,
                user_id=manager.id,
                type="document",
                title=f"✅ Документ {doc.doc_number} проведён",
                message=f"Документ {doc.doc_number} проведён пользователем {current_user.full_name or current_user.login}",
                link="/documents"
            )
        
        return doc
        
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        print(f"🔴 COMPLETE ERROR:\n{str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

# ============================================
# 🔥 ЭКСПОРТ В EXCEL
# ============================================

@router.get("/{doc_id}/export-excel")
def export_document_excel(
    doc_id: int,
    db: Session = Depends(get_db),
    current_user: User = require_worker
):
    """Экспорт документа в Excel с указанием ячеек"""
    try:
        if not XLSX_SUPPORTED:
            raise HTTPException(500, "openpyxl не установлен. Установите: pip install openpyxl")
        
        doc = db.query(WarehouseDocument).filter(WarehouseDocument.id == doc_id).first()
        if not doc:
            raise HTTPException(404, "Документ не найден")
        
        items = db.query(DocumentItem).join(Product).filter(
            DocumentItem.doc_id == doc_id
        ).all()
        
        if not items:
            raise HTTPException(400, "Документ пуст")
        
        # Создаём Excel-книгу
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Document_{doc.doc_number}"
        
        # Стили
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 🔥 ЗАГОЛОВОК ДОКУМЕНТА (на английском для совместимости)
        ws.merge_cells('A1:F1')
        ws['A1'] = f"Document: {doc.doc_number}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # Информация о документе
        doc_type_labels = {
            'receive': 'Receiving',
            'ship': 'Shipping',
            'transfer': 'Transfer',
            'adjust': 'Adjustment'
        }
        
        ws['A2'] = "Document Type:"
        ws['B2'] = doc_type_labels.get(str(doc.type), str(doc.type))
        ws['A3'] = "Status:"
        ws['B3'] = "Completed" if doc.status == "completed" else "Draft"
        ws['A4'] = "Created:"
        ws['B4'] = doc.created_at.strftime("%d.%m.%Y %H:%M") if doc.created_at else "-"
        ws['A5'] = "Comment:"
        ws['B5'] = doc.comment or "-"
        
        # 🔥 ЗАГОЛОВКИ ТАБЛИЦЫ
        headers = ["#", "SKU", "Product Name", "Quantity", "From Cell", "To Cell"]
        row = 7
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        row += 1
        
        # 🔥 ДАННЫЕ
        for idx, item in enumerate(items, 1):
            from_cell = db.query(Cell).filter(Cell.id == item.from_cell_id).first()
            to_cell = db.query(Cell).filter(Cell.id == item.to_cell_id).first()
            
            ws.cell(row=row, column=1, value=idx).border = border
            ws.cell(row=row, column=2, value=item.product.sku if item.product else "-").border = border
            ws.cell(row=row, column=3, value=item.product.name if item.product else f"Product #{item.product_id}").border = border
            ws.cell(row=row, column=4, value=item.quantity).border = border
            ws.cell(row=row, column=5, value=from_cell.code if from_cell else "-").border = border
            ws.cell(row=row, column=6, value=to_cell.code if to_cell else "-").border = border
            
            # Выравнивание
            for col in range(1, 7):
                ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")
            
            row += 1
        
        # 🔥 ИТОГО
        if row > 8:
            total_row = row
            ws.cell(row=total_row, column=2, value="TOTAL:").font = Font(bold=True)
            ws.cell(row=total_row, column=4, value=f"=SUM(D8:D{total_row-1})").font = Font(bold=True)
        
        # Автоширина колонок
        for col in range(1, 7):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 20
        
        # Сохраняем в буфер
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 🔥 ИСПРАВЛЕНО: filename только на английском
        filename = f"Document_{doc.doc_number}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"🔴 EXPORT EXCEL ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(500, f"Ошибка экспорта Excel: {str(e)}")


# ============================================
# 🔥 ЭКСПОРТ В CSV
# ============================================

@router.get("/{doc_id}/export-csv")
def export_document_csv(
    doc_id: int,
    db: Session = Depends(get_db),
    current_user: User = require_worker
):
    """Экспорт документа в CSV"""
    try:
        doc = db.query(WarehouseDocument).filter(WarehouseDocument.id == doc_id).first()
        if not doc:
            raise HTTPException(404, "Документ не найден")
        
        items = db.query(DocumentItem).join(Product).filter(
            DocumentItem.doc_id == doc_id
        ).all()
        
        if not items:
            raise HTTPException(400, "Документ пуст")
        
        # 🔥 ИСПРАВЛЕНО: используем BytesIO вместо StringIO
        output = io.BytesIO()
        
        # Записываем с BOM для Excel
        output.write('\ufeff'.encode('utf-8'))
        
        # Заголовок документа
        output.write(f"Document: {doc.doc_number}\n".encode('utf-8'))
        output.write(f"Type: {doc_type_label(doc.type)}\n".encode('utf-8'))
        output.write(f"Status: {'Completed' if doc.status == 'completed' else 'Draft'}\n".encode('utf-8'))
        output.write(f"Created: {doc.created_at.strftime('%d.%m.%Y %H:%M') if doc.created_at else '-'}\n".encode('utf-8'))
        output.write(f"Comment: {doc.comment or '-'}\n\n".encode('utf-8'))
        
        # Заголовки таблицы
        output.write("№;SKU;Product Name;Quantity;From Cell;To Cell\n".encode('utf-8'))
        
        # Данные
        for idx, item in enumerate(items, 1):
            from_cell = db.query(Cell).filter(Cell.id == item.from_cell_id).first()
            to_cell = db.query(Cell).filter(Cell.id == item.to_cell_id).first()
            
            row = f"{idx};{item.product.sku if item.product else '-'};{item.product.name if item.product else f'Product #{item.product_id}'};{item.quantity};{from_cell.code if from_cell else '-'};{to_cell.code if to_cell else '-'}\n"
            output.write(row.encode('utf-8'))
        
        # Итого
        total_quantity = sum(item.quantity for item in items)
        output.write(f"\nTOTAL;;;{total_quantity};;\n".encode('utf-8'))
        
        output.seek(0)
        
        # 🔥 ИСПРАВЛЕНО: filename только на английском
        filename = f"Document_{doc.doc_number}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        return Response(
            content=output.getvalue(),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"🔴 EXPORT CSV ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(500, f"Ошибка экспорта CSV: {str(e)}")

# ============================================
# ОСТАЛЬНЫЕ ЭНДПОИНТЫ
# ============================================

@router.put("/{doc_id}", response_model=DocumentResponse)
def update_document(
    doc_id: int,
    data: DocumentCreate,
    db: Session = Depends(get_db),
    current_user: User = require_manager
):
    try:
        doc = db.query(WarehouseDocument).filter(WarehouseDocument.id == doc_id).first()
        if not doc:
            raise HTTPException(404, "Документ не найден")
        if doc.status != DocStatus.DRAFT:
            raise HTTPException(400, "Можно редактировать только черновики")
        
        doc.type = data.type
        doc.comment = data.comment
        
        old_items = db.query(DocumentItem).filter(DocumentItem.doc_id == doc_id).all()
        for item in old_items:
            db.delete(item)
        db.flush()
        
        for item_data in data.items:
            db.add(DocumentItem(
                doc_id=doc.id,
                product_id=item_data.product_id,
                quantity=item_data.quantity,
                from_cell_id=item_data.from_cell_id,
                to_cell_id=item_data.to_cell_id
            ))
        
        log_action(db, current_user, "UPDATE", "document", doc_id,
                   new_value={"doc_number": doc.doc_number, "type": doc.type})
        
        db.commit()
        db.refresh(doc)
        
        # Уведомление
        workers = db.query(User).filter(User.role.in_(['admin', 'warehouse_manager', 'warehouse_worker'])).all()
        for worker in workers:
            create_notification(
                db=db,
                user_id=worker.id,
                type="document",
                title=f"📄 Документ {doc.doc_number} обновлён",
                message=f"Документ {doc.doc_number} обновлён",
                link="/documents"
            )
        return doc
    except Exception as e:
        db.rollback()
        raise HTTPException(500, str(e))

@router.patch("/{doc_id}/status", response_model=dict)
def update_document_status(
    doc_id: int,
    data: dict,
    db: Session = Depends(get_db),
    current_user: User = require_manager
):
    try:
        doc = db.query(WarehouseDocument).filter(WarehouseDocument.id == doc_id).first()
        if not doc:
            raise HTTPException(404, "Документ не найден")
        
        new_status = data.get("status")
        old_status = doc.status
        
        if new_status == "cancelled":
            if doc.status in [DocStatus.COMPLETED]:
                raise HTTPException(400, "Нельзя отменить проведённый документ")
            
            doc.status = DocStatus.CANCELLED
            log_action(db, current_user, "CANCEL", "document", doc_id,
                       old_value={"status": old_status},
                       new_value={"status": "cancelled"})
            
            # Уведомление
            workers = db.query(User).filter(User.role.in_(['admin', 'warehouse_manager', 'warehouse_worker'])).all()
            for worker in workers:
                create_notification(
                    db=db,
                    user_id=worker.id,
                    type="document",
                    title=f"❌ Документ {doc.doc_number} отменён",
                    message=f"Документ {doc.doc_number} отменён",
                    link="/documents"
                )
        elif new_status == "draft":
            doc.status = DocStatus.DRAFT
        elif new_status == "in_progress":
            doc.status = DocStatus.IN_PROGRESS
        else:
            raise HTTPException(400, f"Неизвестный статус: {new_status}")
        
        db.commit()
        return {"message": f"Статус изменён на {new_status}"}
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(500, str(e))

@router.delete("/{doc_id}")
def delete_document(
    doc_id: int,
    db: Session = Depends(get_db),
    current_user: User = require_manager
):
    try:
        doc = db.query(WarehouseDocument).filter(WarehouseDocument.id == doc_id).first()
        if not doc:
            raise HTTPException(404, "Документ не найден")
        if doc.status != DocStatus.DRAFT:
            raise HTTPException(400, "Можно удалять только черновики")
        
        items = db.query(DocumentItem).filter(DocumentItem.doc_id == doc_id).all()
        for item in items:
            db.delete(item)
        
        db.delete(doc)
        
        log_action(db, current_user, "DELETE", "document", doc_id,
                   new_value={"doc_number": doc.doc_number})
        
        db.commit()
        
        # Уведомление
        workers = db.query(User).filter(User.role.in_(['admin', 'warehouse_manager', 'warehouse_worker'])).all()
        for worker in workers:
            create_notification(
                db=db,
                user_id=worker.id,
                type="document",
                title=f"🗑️ Документ {doc.doc_number} удалён",
                message=f"Документ {doc.doc_number} удалён",
                link="/documents"
            )
        return {"message": "Документ удалён"}
    except Exception as e:
        db.rollback()
        raise HTTPException(500, str(e))