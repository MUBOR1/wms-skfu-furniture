from fastapi import APIRouter, Depends, HTTPException, Query, Response
from sqlalchemy.orm import Session
from sqlalchemy import func, select, cast, Date, or_
from database import get_db
from core.permissions import require_worker, require_manager
from models.user import User
from models.product import Product
from models.stock import Stock
from models.document import WarehouseDocument, DocStatus, DocumentItem
from models.order import Order, OrderStatus, OrderItem
from models.cell import Cell
from routers.notifications import create_notification
from datetime import datetime, timedelta
from typing import Optional
import io
import csv

# 🔥 ИМПОРТЫ ДЛЯ EXCEL
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    XLSX_SUPPORTED = True
except ImportError:
    XLSX_SUPPORTED = False
    print("⚠️ openpyxl не установлен. Excel экспорт будет недоступен.")

last_critical_notification = {}

router = APIRouter(prefix="/api/analytics", tags=["Аналитика"])

# ============================================
# 🔥 СУЩЕСТВУЮЩИЕ ЭНДПОИНТЫ (без изменений)
# ============================================

@router.get("/stock-report")
def get_stock_report(db: Session = Depends(get_db), current_user: User = require_worker):
    """Отчёт по складским остаткам с группировкой по товарам"""
    
    results = db.query(
        Product.id,
        Product.sku,
        Product.name,
        Product.category,
        Product.purchase_price,
        Product.sale_price,
        Product.min_stock,
        Product.max_stock,
        func.coalesce(func.sum(Stock.quantity), 0).label('total_qty')
    ).outerjoin(
        Stock, Stock.product_id == Product.id
    ).filter(
        Product.is_active == True
    ).group_by(
        Product.id
    ).all()
    
    result_list = []
    critical_products = []
    
    for r in results:
        qty = r.total_qty or 0
        status = "normal"
        if qty == 0:
            status = "critical"
            critical_products.append({"id": r.id, "name": r.name or "Без названия"})
        elif r.min_stock and qty < r.min_stock:
            status = "low"
            critical_products.append({"id": r.id, "name": r.name or "Без названия"})
        elif r.max_stock and qty > r.max_stock:
            status = "overstock"
        
        result_list.append({
            "id": r.id,
            "sku": r.sku or "—",
            "name": r.name or "Без названия",
            "category": r.category or "—",
            "purchase_price": float(r.purchase_price) if r.purchase_price else 0.0,
            "sale_price": float(r.sale_price) if r.sale_price else 0.0,
            "quantity": qty,
            "min_stock": r.min_stock or 0,
            "max_stock": r.max_stock or 0,
            "status": status
        })
    
    # 🔥 УВЕДОМЛЕНИЕ О КРИТИЧЕСКИХ ОСТАТКАХ — НЕ СПАМИТЬ
    if critical_products and current_user.role in ['admin', 'warehouse_manager']:
        global last_critical_notification
        now = datetime.now()
        last_time = last_critical_notification.get('time')
        
        # Проверяем, прошло ли 30 минут с последнего уведомления
        if not last_time or (now - last_time) > timedelta(minutes=30):
            managers = db.query(User).filter(User.role.in_(['admin', 'warehouse_manager'])).all()
            for manager in managers:
                create_notification(
                    db=db,
                    user_id=manager.id,
                    type="alert",
                    title="⚠️ Критические остатки",
                    message=f"{len(critical_products)} товаров требуют пополнения",
                    link="/report"
                )
            last_critical_notification['time'] = now
            last_critical_notification['count'] = len(critical_products)
    
    return result_list

@router.get("/stock-details")
def get_stock_details(
    product_id: int = Query(...), 
    db: Session = Depends(get_db), 
    current_user: User = require_worker
):
    """Возвращает список ячеек и количество товара в каждой"""
    rows = db.query(
        Cell.code.label('cell_code'),
        Stock.quantity
    ).join(
        Cell, Cell.id == Stock.cell_id
    ).filter(
        Stock.product_id == product_id,
        Stock.quantity > 0
    ).order_by(Cell.code).all()

    return [{"cell_code": r.cell_code, "quantity": r.quantity} for r in rows]

@router.get("/dashboard-stats")
def get_dashboard_stats(days: int = 30, db: Session = Depends(get_db), current_user: User = require_worker):
    """Статистика для дашборда — С ГРАФИКАМИ"""
    
    total_stock = db.query(func.coalesce(func.sum(Stock.quantity), 0)).scalar() or 0
    total_products = db.query(func.count(Product.id)).filter(Product.is_active == True).scalar() or 0
    
    subquery = db.query(
        Product.id
    ).join(
        Stock, Stock.product_id == Product.id, isouter=True
    ).filter(
        Product.is_active == True
    ).group_by(
        Product.id, Product.min_stock
    ).having(
        (Product.min_stock > 0) & 
        (func.coalesce(func.sum(Stock.quantity), 0) < Product.min_stock)
    ).subquery()
    
    low_stock = db.query(func.count(subquery.c.id)).scalar() or 0
    
    date_from = datetime.now() - timedelta(days=days)
    daily_turnover = db.query(
        cast(WarehouseDocument.created_at, Date).label('date'),
        func.count(WarehouseDocument.id).label('count')
    ).filter(
        WarehouseDocument.created_at >= date_from,
        WarehouseDocument.status == DocStatus.COMPLETED
    ).group_by(
        cast(WarehouseDocument.created_at, Date)
    ).order_by(
        cast(WarehouseDocument.created_at, Date)
    ).all()
    
    daily_turnover_list = [{"date": str(d.date), "count": d.count} for d in daily_turnover]
    
    top_products = db.query(
        Product.id,
        Product.name,
        func.sum(DocumentItem.quantity).label('total_qty')
    ).join(
        DocumentItem, DocumentItem.product_id == Product.id
    ).join(
        WarehouseDocument, WarehouseDocument.id == DocumentItem.doc_id
    ).filter(
        WarehouseDocument.status == DocStatus.COMPLETED,
        WarehouseDocument.type == "ship",
        WarehouseDocument.created_at >= date_from,
        Product.is_active == True
    ).group_by(
        Product.id, Product.name
    ).order_by(
        func.sum(DocumentItem.quantity).desc()
    ).limit(5).all()
    
    top_products_list = [{"name": p.name, "total_qty": p.total_qty} for p in top_products]
    
    orders_by_status = db.query(
        Order.status,
        func.count(Order.id).label('count')
    ).group_by(Order.status).all()
    
    return {
        "summary": {
            "total_products": total_products,
            "total_stock": total_stock,
            "low_stock": low_stock
        },
        "daily_turnover": daily_turnover_list,
        "top_products": top_products_list,
        "order_statuses": [
            {"status": str(s.status.value) if hasattr(s.status, 'value') else str(s.status), "count": s.count} 
            for s in orders_by_status
        ]
    }


# ============================================
# 🔥 НОВЫЕ ОТЧЁТЫ
# ============================================

@router.get("/turnover-report")
def get_turnover_report(
    period: str = Query("30", description="Период в днях"),
    category: Optional[str] = Query(None, description="Категория товара"),
    db: Session = Depends(get_db),
    current_user: User = require_worker
):
    """Отчёт по оборачиваемости товаров"""
    
    days = int(period) if period.isdigit() else 30
    date_from = datetime.now() - timedelta(days=days)
    
    # 🔥 ПРОВЕРЬТЕ: какие статусы есть в вашей модели OrderStatus
    # Возможные варианты: 'completed', 'delivered', 'shipped', 'finished'
    completed_statuses = ['completed', 'delivered', 'shipped']
    
    query = db.query(
        Product.id,
        Product.sku,
        Product.name,
        Product.category,
        Product.sale_price,
        func.coalesce(func.sum(Stock.quantity), 0).label('stock_qty'),
        func.coalesce(func.sum(OrderItem.quantity), 0).label('sold_qty')
    ).outerjoin(
        Stock, Stock.product_id == Product.id
    ).outerjoin(
        OrderItem, OrderItem.product_id == Product.id
    ).outerjoin(
        Order, Order.id == OrderItem.order_id
    ).filter(
        Product.is_active == True,
        (Order.status.in_(completed_statuses)) | (Order.status.is_(None))
    )
    
    if category:
        query = query.filter(Product.category == category)
    
    if days > 0:
        query = query.filter(
            (Order.created_at >= date_from) | (Order.created_at.is_(None))
        )
    
    query = query.group_by(Product.id, Product.sku, Product.name, Product.category, Product.sale_price)
    results = query.all()
    
    result_list = []
    for r in results:
        stock_qty = r.stock_qty or 0
        sold_qty = r.sold_qty or 0
        turnover = round(sold_qty / stock_qty, 2) if stock_qty > 0 else 0
        
        result_list.append({
            "id": r.id,
            "sku": r.sku or "—",
            "name": r.name or "Без названия",
            "category": r.category or "—",
            "sale_price": float(r.sale_price) if r.sale_price else 0,
            "stock_qty": stock_qty,
            "sold_qty": sold_qty,
            "turnover": turnover
        })
    
    return result_list


@router.get("/critical-report")
def get_critical_report(
    category: Optional[str] = Query(None, description="Категория товара"),
    db: Session = Depends(get_db),
    current_user: User = require_worker
):
    """Отчёт по критическим остаткам"""
    
    query = db.query(
        Product.id,
        Product.sku,
        Product.name,
        Product.category,
        Product.min_stock,
        Product.max_stock,
        func.coalesce(func.sum(Stock.quantity), 0).label('stock_qty')
    ).outerjoin(
        Stock, Stock.product_id == Product.id
    ).filter(
        Product.is_active == True
    )
    
    if category:
        query = query.filter(Product.category == category)
    
    query = query.group_by(Product.id, Product.sku, Product.name, Product.category, Product.min_stock, Product.max_stock)
    query = query.having(
        (func.coalesce(func.sum(Stock.quantity), 0) < Product.min_stock) |
        (func.coalesce(func.sum(Stock.quantity), 0) == 0) |
        (func.coalesce(func.sum(Stock.quantity), 0) > Product.max_stock)
    )
    
    results = query.all()
    
    result_list = []
    for r in results:
        stock_qty = r.stock_qty or 0
        status = "critical"
        if r.min_stock and stock_qty > 0 and stock_qty < r.min_stock:
            status = "low"
        elif r.max_stock and stock_qty > r.max_stock:
            status = "overstock"
        elif stock_qty == 0:
            status = "critical"
        
        result_list.append({
            "id": r.id,
            "sku": r.sku or "—",
            "name": r.name or "Без названия",
            "category": r.category or "—",
            "stock_qty": stock_qty,
            "min_stock": r.min_stock or 0,
            "max_stock": r.max_stock or 0,
            "status": status
        })
    
    return result_list


@router.get("/value-report")
def get_value_report(
    category: Optional[str] = Query(None, description="Категория товара"),
    db: Session = Depends(get_db),
    current_user: User = require_worker
):
    """Отчёт по стоимости запасов"""
    
    query = db.query(
        Product.id,
        Product.sku,
        Product.name,
        Product.category,
        Product.purchase_price,
        Product.sale_price,
        func.coalesce(func.sum(Stock.quantity), 0).label('stock_qty')
    ).outerjoin(
        Stock, Stock.product_id == Product.id
    ).filter(
        Product.is_active == True
    )
    
    if category:
        query = query.filter(Product.category == category)
    
    query = query.group_by(Product.id, Product.sku, Product.name, Product.category, Product.purchase_price, Product.sale_price)
    results = query.all()
    
    total_value = 0
    result_list = []
    
    for r in results:
        stock_qty = r.stock_qty or 0
        purchase_price = float(r.purchase_price) if r.purchase_price else 0
        sale_price = float(r.sale_price) if r.sale_price else 0
        
        purchase_value = stock_qty * purchase_price
        sale_value = stock_qty * sale_price
        total_value += purchase_value
        
        result_list.append({
            "id": r.id,
            "sku": r.sku or "—",
            "name": r.name or "Без названия",
            "category": r.category or "—",
            "stock_qty": stock_qty,
            "purchase_price": round(purchase_price, 2),
            "sale_price": round(sale_price, 2),
            "purchase_value": round(purchase_value, 2),
            "sale_value": round(sale_value, 2)
        })
    
    return {
        "items": result_list,
        "total_value": round(total_value, 2)
    }


# ============================================
# 🔥 ЭКСПОРТ ОТЧЁТОВ В EXCEL
# ============================================

@router.get("/export-excel")
def export_report_excel(
    report_type: str = Query(..., description="Тип отчёта: stock, turnover, critical, value"),
    period: Optional[str] = Query("30", description="Период в днях"),
    category: Optional[str] = Query(None, description="Категория товара"),
    db: Session = Depends(get_db),
    current_user: User = require_worker
):
    """Экспорт отчёта в Excel"""
    try:
        if not XLSX_SUPPORTED:
            raise HTTPException(500, "openpyxl не установлен. Установите: pip install openpyxl")
        
        # Получаем данные в зависимости от типа отчёта
        if report_type == "stock":
            data = get_stock_report(db, current_user)
            title = "Отчёт по остаткам"
            headers = ["SKU", "Наименование", "Категория", "Остаток", "Мин.", "Макс.", "Статус"]
            rows = []
            for item in data:
                status_map = {
                    "critical": "Нет в наличии",
                    "low": "Мало",
                    "overstock": "Переизбыток",
                    "normal": "Норма"
                }
                rows.append([
                    item.get("sku", "—"),
                    item.get("name", "—"),
                    item.get("category", "—"),
                    item.get("quantity", 0),
                    item.get("min_stock", 0),
                    item.get("max_stock", 0),
                    status_map.get(item.get("status", "normal"), "Норма")
                ])
            
        elif report_type == "turnover":
            data = get_turnover_report(period, category, db, current_user)
            title = "Отчёт по оборачиваемости"
            headers = ["SKU", "Наименование", "Категория", "Цена", "Остаток", "Продано", "Оборачиваемость"]
            rows = []
            for item in data:
                rows.append([
                    item.get("sku", "—"),
                    item.get("name", "—"),
                    item.get("category", "—"),
                    item.get("sale_price", 0),
                    item.get("stock_qty", 0),
                    item.get("sold_qty", 0),
                    item.get("turnover", 0)
                ])
            
        elif report_type == "critical":
            data = get_critical_report(category, db, current_user)
            title = "Отчёт по критическим остаткам"
            headers = ["SKU", "Наименование", "Категория", "Остаток", "Мин.", "Макс.", "Статус"]
            rows = []
            for item in data:
                status_map = {
                    "critical": "Нет в наличии",
                    "low": "Мало",
                    "overstock": "Переизбыток"
                }
                rows.append([
                    item.get("sku", "—"),
                    item.get("name", "—"),
                    item.get("category", "—"),
                    item.get("stock_qty", 0),
                    item.get("min_stock", 0),
                    item.get("max_stock", 0),
                    status_map.get(item.get("status", "critical"), "Нет в наличии")
                ])
            
        elif report_type == "value":
            data = get_value_report(category, db, current_user)
            title = "Отчёт по стоимости запасов"
            headers = ["SKU", "Наименование", "Категория", "Остаток", "Закупка", "Продажа", "Стоимость закупки", "Стоимость продажи"]
            rows = []
            for item in data.get("items", []):
                rows.append([
                    item.get("sku", "—"),
                    item.get("name", "—"),
                    item.get("category", "—"),
                    item.get("stock_qty", 0),
                    item.get("purchase_price", 0),
                    item.get("sale_price", 0),
                    item.get("purchase_value", 0),
                    item.get("sale_value", 0)
                ])
            
            # Добавляем итоговую строку
            total_value = data.get("total_value", 0)
            rows.append(["", "", "", "", "", "ИТОГО:", total_value, ""])
        else:
            raise HTTPException(400, "Неизвестный тип отчёта")
        
        # Создаём Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title[:31]
        
        # Стили
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Заголовок
        ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # Дата и пользователь
        ws['A2'] = f"Дата: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ws['A3'] = f"Пользователь: {current_user.full_name or current_user.login}"
        
        # Заголовки таблицы
        row = 5
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        row += 1
        
        # Данные
        for data_row in rows:
            for col, value in enumerate(data_row, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
            row += 1
        
        # Автоширина
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 20
        
        # Сохраняем в буфер
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 🔥 ИСПРАВЛЕНО: имя файла только на английском
        type_names = {
            'stock': 'Ostatki',
            'turnover': 'Oborachivaemost',
            'critical': 'Kriticheskie',
            'value': 'Stoimost'
        }
        filename = f"Report_{type_names.get(report_type, 'report')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"🔴 EXPORT EXCEL ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(500, f"Ошибка экспорта Excel: {str(e)}")